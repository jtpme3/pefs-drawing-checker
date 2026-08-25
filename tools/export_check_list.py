"""Export the full list of checks the tool runs, scraped from the source.

Usage:  python tools/export_check_list.py [output.xlsx]

The list is built from the check definitions in ``drawing_checker/`` rather
than typed out by hand, so it cannot drift from what actually runs. Each row
carries the code, the check wording, its group, the checklist rows it came
from (``ARN10-G-LIST-0135 Rev A``), and the module that implements it --
open that module to read exactly how the check decides.
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

PACKAGE = Path(__file__).resolve().parent.parent / "drawing_checker"

MODULES = [
    "checks.py", "content_checks.py", "specs.py", "geometry_checks.py",
    "notes.py", "client_checks.py", "maop.py", "flow_checks.py",
    "dxf_checks.py",
]

CODE = r"[A-Z]{2,5}-\d{2}"
STRING = r'"[^"]*"(?:\s*\n\s*"[^"]*")*'          # possibly-continued literal

# add("CODE", "group", "title"...   /  Result("CODE", "group", "title"...
LITERAL = re.compile(
    rf'(?:add\(|Result\()\s*\n?\s*"(?P<code>{CODE})",\s*"(?P<group>[^"]+)",'
    rf'\s*\n?\s*(?P<title>{STRING})',
    re.MULTILINE,
)
# add("CODE", "group", title, ...  -- title held in a variable defined above
VARIABLE = re.compile(
    rf'(?:add\(|Result\()\s*\n?\s*"(?P<code>{CODE})",\s*"(?P<group>[^"]+)",'
    r'\s*\n?\s*(?P<var>[a-z_]+)\s*,',
    re.MULTILINE,
)
# ref strings look like "Rows 20, 21" / "Row 82" / "n/a - tool coverage"
REF = re.compile(r'"((?:Rows?\s[\d,\s]+|n/a[^"]*))"')


def _join(literal: str) -> str:
    return " ".join(re.findall(r'"([^"]*)"', literal))


def scrape() -> list[dict]:
    found: dict[str, dict] = {}
    for name in MODULES:
        src = (PACKAGE / name).read_text(encoding="utf-8")
        for match in LITERAL.finditer(src):
            _record(found, src, name, match, _join(match["title"]))
        for match in VARIABLE.finditer(src):
            if match["code"] in found:
                continue
            # nearest `var = "..."` above the call
            var = match["var"]
            defs = list(re.finditer(
                rf'^\s*{var}\s*=\s*\(?\s*({STRING})', src[: match.start()],
                re.MULTILINE))
            if defs:
                _record(found, src, name, match, _join(defs[-1].group(1)))
    return [found[c] for c in sorted(found)]


def _record(found, src, module, match, title) -> None:
    code = match["code"]
    if code in found:
        return
    # Checklist ref: either a literal inside this call's remaining arguments,
    # or the bare identifier ``ref`` (defined as ``ref = "Rows ..."`` above
    # the call) -- whichever appears first in the call tail.
    tail = src[match.end(): match.end() + 800]
    literal = REF.search(tail)
    ident = re.search(r"[,(]\s*\n?\s*ref\s*[,)]", tail)
    ref = ""
    if literal and (not ident or literal.start() < ident.start()):
        ref = literal.group(1)
    elif ident:
        defs = re.findall(r'^\s*ref\s*=\s*"([^"]+)"', src[: match.start()],
                          re.MULTILINE)
        ref = defs[-1] if defs else ""
    found[code] = {
        "code": code,
        "group": match["group"],
        "title": title,
        "ref": ref,
        "module": f"drawing_checker/{module}",
    }


def main() -> int:
    rows = scrape()
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("check list.xlsx")

    import openpyxl
    from openpyxl.styles import Font

    # Jordan writes review comments into a Comments column of the exported
    # workbook. Regenerating must never wipe them, so carry any existing
    # Comments across by check code before overwriting the file.
    comments: dict[str, str] = {}
    uncoded: list[str] = []  # comment rows with no check code (new requests)
    if out.exists():
        old = openpyxl.load_workbook(out)
        ws = old.active
        header = [c.value for c in ws[1]]
        if "Comments" in header:
            col = header.index("Comments")
            for r in ws.iter_rows(min_row=2, values_only=True):
                if col < len(r) and r[col]:
                    if r[0]:
                        comments[str(r[0])] = str(r[col])
                    else:
                        uncoded.append(str(r[col]))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Checks"
    headers = ["Code", "Check", "Group", "Checklist rows", "Implemented in",
               "Comments"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([row["code"], row["title"], row["group"], row["ref"],
                   row["module"], comments.get(row["code"], "")])
    for text in uncoded:
        ws.append(["", "", "", "", "", text])
    for col, width in zip("ABCDEF", (10, 62, 16, 22, 34, 60)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    wb.save(out)
    if comments:
        print(f"(carried {len(comments)} existing comment(s) across)")

    for row in rows:
        print(f"{row['code']:9}| {row['group']:15}| {row['title']}")
    print(f"\n{len(rows)} checks -> {out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
