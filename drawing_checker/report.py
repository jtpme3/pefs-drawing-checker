"""Write the check results to an Excel workbook.

Three sheets:
  Summary   one row per drawing, worst-status first
  Findings  one row per thing to look at (FAIL / REVIEW / NO DATA)
  Matrix    checks down, drawings across -- the same shape as the manual
            PEFS drawing checklist, so results can be read the familiar way
"""

from __future__ import annotations

from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .checks import FAIL, NA, NO_DATA, PASS, REVIEW, SKIPPED, Result

FILLS = {
    PASS: PatternFill("solid", fgColor="C6EFCE"),
    FAIL: PatternFill("solid", fgColor="FFC7CE"),
    REVIEW: PatternFill("solid", fgColor="FFEB9C"),
    NO_DATA: PatternFill("solid", fgColor="D9D9D9"),
    NA: PatternFill("solid", fgColor="F2F2F2"),
    SKIPPED: PatternFill("solid", fgColor="EDEDED"),
}
FONTS = {
    PASS: Font(color="006100"),
    FAIL: Font(color="9C0006", bold=True),
    REVIEW: Font(color="9C5700"),
    NO_DATA: Font(color="404040"),
    NA: Font(color="808080"),
    SKIPPED: Font(color="A6A6A6", italic=True),
}
HEADER = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="44546A")
TITLE = Font(bold=True, size=14)
THIN = Border(*[Side("thin", color="BFBFBF")] * 4)

# Worst status first when sorting.
RANK = {FAIL: 0, REVIEW: 1, NO_DATA: 2, PASS: 3, NA: 4, SKIPPED: 5}


@dataclass
class SheetReport:
    """One checked page, ready to be written out."""

    label: str          # "Q-4300-10-DF-193_1_IFC.pdf" or "... (p3)"
    filename: str       # source PDF name, without the page suffix
    drawing_number: str
    revision: str
    results: list[Result] = field(default_factory=list)

    @property
    def skipped(self) -> bool:
        return any(r.status == SKIPPED for r in self.results)

    @property
    def worst(self) -> str:
        if self.skipped:
            return SKIPPED
        return min(
            (r.status for r in self.results), key=lambda s: RANK.get(s, 9), default=NA
        )

    def count(self, status: str) -> int:
        return sum(1 for r in self.results if r.status == status)


def _header_row(ws, row: int, values: list[str]) -> None:
    for col, value in enumerate(values, start=1):
        c = ws.cell(row, col, value)
        c.font = HEADER
        c.fill = HEADER_FILL
        c.alignment = Alignment(vertical="center", wrap_text=True)


def _status_cell(ws, row: int, col: int, status: str):
    c = ws.cell(row, col, status)
    c.fill = FILLS.get(status, FILLS[NA])
    c.font = FONTS.get(status, FONTS[NA])
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = THIN
    return c


def write_report(sheets: list[SheetReport], out_path: str | Path) -> Path:
    out_path = Path(out_path)
    checked = [s for s in sheets if not s.skipped]
    skipped = [s for s in sheets if s.skipped]

    wb = Workbook()
    _write_summary(wb.active, checked, skipped)
    _write_findings(wb.create_sheet("Findings"), checked)
    _write_matrix(wb.create_sheet("Matrix"), checked)

    wb.save(out_path)
    return out_path


def _write_summary(ws, checked: list[SheetReport], skipped: list[SheetReport]) -> None:
    ws.title = "Summary"
    ws["A1"] = "PEFS drawing check - summary"
    ws["A1"].font = TITLE
    ws["A2"] = (
        f"Generated {datetime.now():%Y-%m-%d %H:%M}  -  checks: drawing frame "
        "and identity (slice 1)"
    )
    ws["A3"] = (
        "FAIL = objectively wrong, fix before issue.  REVIEW = needs an "
        "engineer's eye.  NO DATA = the tool could not read the field."
    )

    _header_row(ws, 5, [
        "Drawing file", "Drawing no.", "Rev", "Overall",
        "Fail", "Review", "No data", "Pass", "N/A", "Top issue",
    ])

    row = 6
    for s in sorted(checked, key=lambda s: (RANK.get(s.worst, 9), s.label)):
        top = next(
            (r for r in sorted(s.results, key=lambda r: RANK.get(r.status, 9))
             if r.status in (FAIL, REVIEW, NO_DATA)),
            None,
        )
        ws.cell(row, 1, s.label)
        ws.cell(row, 2, s.drawing_number)
        ws.cell(row, 3, s.revision)
        _status_cell(ws, row, 4, s.worst)
        for col, status in enumerate([FAIL, REVIEW, NO_DATA, PASS, NA], start=5):
            ws.cell(row, col, s.count(status)).alignment = Alignment(
                horizontal="center"
            )
        ws.cell(row, 10, f"{top.code} {top.check}" if top else "")
        row += 1

    if not checked:
        ws.cell(row, 1, "No Origin drawing sheets were found in the input.")
        row += 1

    if skipped:
        row += 1
        c = ws.cell(row, 1, f"Not checked - not Origin drawing sheets ({len(skipped)} page(s))")
        c.font = Font(bold=True, italic=True, color="808080")
        row += 1
        for name, pages in sorted(Counter(s.filename for s in skipped).items()):
            ws.cell(row, 1, name).font = FONTS[SKIPPED]
            ws.cell(row, 2, f"{pages} page(s)").font = FONTS[SKIPPED]
            row += 1

    for col, width in enumerate([52, 22, 6, 10, 7, 8, 9, 7, 7, 46], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A6"


def _write_findings(ws, checked: list[SheetReport]) -> None:
    ws["A1"] = "Findings - everything that is not a clean pass"
    ws["A1"].font = TITLE
    _header_row(ws, 3, [
        "Drawing file", "Drawing no.", "Rev", "Status", "Check",
        "What was found", "Checklist row",
    ])

    row = 4
    flagged = [
        (s, r) for s in checked for r in s.results
        if r.status in (FAIL, REVIEW, NO_DATA)
    ]
    for s, r in sorted(
        flagged, key=lambda f: (RANK.get(f[1].status, 9), f[0].label, f[1].code)
    ):
        ws.cell(row, 1, s.label)
        ws.cell(row, 2, s.drawing_number)
        ws.cell(row, 3, s.revision)
        _status_cell(ws, row, 4, r.status)
        ws.cell(row, 5, f"{r.code}  {r.check}")
        ws.cell(row, 6, r.detail).alignment = Alignment(
            wrap_text=True, vertical="top"
        )
        ws.cell(row, 7, r.checklist_ref)
        row += 1

    if row == 4:
        ws.cell(4, 1, "No findings - every check passed on every drawing.")

    for col, width in enumerate([46, 22, 6, 10, 46, 88, 14], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A4"


def _write_matrix(ws, checked: list[SheetReport]) -> None:
    """Checks down the page, drawings across -- the manual checklist layout."""
    ws["A1"] = "Check matrix"
    ws["A1"].font = TITLE

    codes: list[tuple[str, str, str, str]] = []
    seen = set()
    for s in checked:
        for r in s.results:
            if r.code not in seen:
                seen.add(r.code)
                codes.append((r.code, r.group, r.check, r.checklist_ref))

    _header_row(ws, 3, ["Code", "Group", "Check", "Checklist row"])
    for col, s in enumerate(checked, start=5):
        c = ws.cell(3, col, s.drawing_number or s.label)
        c.font = HEADER
        c.fill = HEADER_FILL
        c.alignment = Alignment(textRotation=90, horizontal="center", vertical="bottom")
        ws.column_dimensions[get_column_letter(col)].width = 11

    by_file = [{r.code: r.status for r in s.results} for s in checked]
    for row, (code, group, check, ref) in enumerate(codes, start=4):
        ws.cell(row, 1, code)
        ws.cell(row, 2, group)
        ws.cell(row, 3, check)
        ws.cell(row, 4, ref)
        for col, statuses in enumerate(by_file, start=5):
            _status_cell(ws, row, col, statuses.get(code, NA))

    for col, width in enumerate([9, 16, 46, 14], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[3].height = 110
    ws.freeze_panes = "E4"
