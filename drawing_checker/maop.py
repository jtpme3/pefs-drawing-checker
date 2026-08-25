"""Cross-check drawn line numbers against the MAOP review calculation.

The MAOP review (``ARN10-P-CALC-0029``) is the process authority for what each
corridor's line number should be: it fixes the size, the SDR and therefore the
pipe spec. This module reads it and compares against what the drawings show.

The trailing project code is deliberately ignored. Jordan changes it to match
the area number on the drawing, so ``...-P151-4255`` on the calc and
``...-P151-4200`` on the drawing are the same line. The sequence field is
ignored too -- it is ``XXXX`` throughout the calc.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path

# Corridor tag as drawn and as listed in the calc, e.g. CMN127-4, PHS045-3.
CORRIDOR = re.compile(r"\b([A-Z]{3}\d{3}-\d+)\b")
LINE_NUMBER = re.compile(
    r"\b(?P<size>\d{2,4})-(?P<service>[A-Z]{2})-(?P<seq>[A-Z0-9]+)"
    r"-(?P<spec>[A-Z]\d{2,3})-(?P<project>\d{4})\b"
)

# Sheet name -> (corridor column, line-number column), 1-based, header on row 9.
SHEETS = {
    "Gas Flowline SDR Check": (1, 17),
    "Water Flowline SDR Check": (1, 19),
}
FIRST_DATA_ROW = 10


@dataclass(frozen=True)
class Expectation:
    corridor: str
    size: str
    service: str
    spec: str
    source_sheet: str

    @property
    def key(self) -> tuple[str, str, str]:
        return (self.size, self.service, self.spec)

    def __str__(self) -> str:
        return f"{self.size}-{self.service}-XXXX-{self.spec}"


@dataclass
class MaopDataset:
    path: Path | None = None
    by_corridor: dict[str, list[Expectation]] = field(default_factory=dict)

    @property
    def loaded(self) -> bool:
        return bool(self.by_corridor)

    def expectations(self, corridor: str) -> list[Expectation]:
        return self.by_corridor.get(corridor.upper(), [])


def load(reference_dir: Path) -> MaopDataset:
    """Read the newest MAOP review calculation in ``reference_dir``."""
    try:
        import openpyxl
    except ImportError:  # pragma: no cover
        return MaopDataset()

    candidates = [
        p for p in reference_dir.rglob("*MAOP*Review*Calculation*.xlsx")
        if not p.name.startswith("~$")
    ]
    if not candidates:
        return MaopDataset()
    path = max(candidates, key=lambda p: p.stat().st_mtime)

    dataset = MaopDataset(path=path)
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:
        return MaopDataset()

    for sheet_name, (corridor_col, line_col) in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in range(FIRST_DATA_ROW, ws.max_row + 1):
            corridor = ws.cell(row, corridor_col).value
            line = ws.cell(row, line_col).value
            if not corridor or not line:
                continue
            match = LINE_NUMBER.search(str(line).strip().upper())
            if not match:
                continue
            dataset.by_corridor.setdefault(
                str(corridor).strip().upper(), []
            ).append(Expectation(
                corridor=str(corridor).strip().upper(),
                size=match["size"],
                service=match["service"],
                spec=match["spec"],
                source_sheet=sheet_name,
            ))
    return dataset


def run_maop_check(sheet, tb, dataset: MaopDataset, add) -> None:
    """Every corridor on the sheet should carry the line number the calc sets."""
    from .checks import FAIL, NA, NO_DATA, PASS
    from .content_checks import body_items

    ref = "Rows 23, 24, 53, 54"
    title = "Line numbers agree with the MAOP review calculation"

    if not dataset.loaded:
        add("MAOP-01", "Piping Spec", title, NO_DATA,
            "No MAOP review calculation found in reference/.", ref)
        return

    items = body_items(sheet)
    blob = " ".join(i.text for i in items)
    corridors = sorted(set(CORRIDOR.findall(blob)))
    drawn = {
        (m["size"], m["service"], m["spec"])
        for m in LINE_NUMBER.finditer(blob)
    }

    covered = [c for c in corridors if dataset.expectations(c)]
    if not covered:
        add("MAOP-01", "Piping Spec", title, NA,
            f"None of this sheet's corridors are in {dataset.path.name}"
            + (f" (drawn: {', '.join(corridors[:8])})" if corridors else ""),
            ref)
        return

    missing: list[str] = []
    flagged: list[str] = []
    for corridor in covered:
        for expectation in dataset.expectations(corridor):
            if expectation.key not in drawn:
                missing.append(
                    f"{corridor} should carry {expectation} "
                    f"({expectation.source_sheet.split()[0].lower()}), which is "
                    "not on this sheet"
                )
                if corridor not in flagged:
                    flagged.append(corridor)

    if missing:
        # Anchor the finding on the drawn corridor labels it is about. Left to
        # the markup pass's text search, the detail's quoted calc filename and
        # corridor fragments latch onto the reference-drawings table and the
        # CADFILE stamp instead of the corridor. A corridor drawn only inside
        # a crossing or tie-in tag (TIP-CMN296-1) still anchors there.
        def names_corridor(text: str, corridor: str) -> bool:
            bare = re.sub(r"^(?:HPC|RC|WC|HVC|PLC|TIP)-", "", text)
            return bare == corridor or bare.startswith(corridor + "-")

        anchors = [
            (i.x0, i.y0, i.x1, i.y1)
            for corridor in flagged
            for i in items
            if names_corridor(i.text.strip(), corridor)
        ]
        add("MAOP-01", "Piping Spec", title, FAIL,
            "; ".join(missing)
            + f". Compared against {dataset.path.name}; sequence number and "
              "project code ignored.",
            ref, anchors=anchors[:8])
    else:
        add("MAOP-01", "Piping Spec", title, PASS,
            f"{len(covered)} corridor(s) in the MAOP review "
            f"({', '.join(covered[:6])}{'...' if len(covered) > 6 else ''}) "
            "all carry the line number it sets.",
            ref)
